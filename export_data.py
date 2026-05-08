"""
export_data.py — Converts SSOT.xlsx → JSON files in data/
Run: python export_data.py
"""
import json, os, sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'SSOT.xlsx')
OUT  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)

def sheet_to_rows(name):
    ws = wb[name]
    headers = [c.value for c in ws[1] if c.value is not None]
    rows = []
    for r in ws.iter_rows(min_row=2, max_col=len(headers)):
        vals = [c.value for c in r]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows

def serialize(obj):
    if isinstance(obj, (datetime, date)):
        return obj.strftime('%Y-%m-%d')
    return obj

def write(name, data):
    path = os.path.join(OUT, f'{name}.json')
    with open(path, 'w') as f:
        json.dump(data, f, indent=2, default=serialize)
    print(f'  {name}.json')

# ---- META ----
meta_rows = sheet_to_rows('META')
meta = {}
for r in meta_rows:
    k, v = r.get('Key'), r.get('Value')
    if k:
        # Try to preserve numeric types
        if isinstance(v, str):
            try: v = float(v) if '.' in v else int(v)
            except: pass
        meta[k] = v
meta['data_date'] = meta.get('data_date', datetime.now().strftime('%Y-%m-%d'))
write('meta', meta)

# ---- Contradictions ----
write('contradictions', sheet_to_rows('Contradictions'))

# ---- Cost_Performance ----
write('cost_performance', sheet_to_rows('Cost_Performance'))

# ---- Actions_Log ----
write('actions_log', sheet_to_rows('Actions_Log'))

# ---- Float_Erosion ----
write('float_erosion', sheet_to_rows('Float_Erosion'))

# ---- Production_Validation ----
write('production_validation', sheet_to_rows('Production_Validation'))

# ---- Productivity_Factor ----
write('productivity_factor', sheet_to_rows('Productivity_Factor'))

# ---- Recovery_Plans ----
write('recovery_plans', sheet_to_rows('Recovery_Plans'))

# ---- Subcontractors ----
write('subcontractors', sheet_to_rows('Subcontractors'))

# ---- Procurement ----
write('procurement', sheet_to_rows('Procurement'))

# ---- Source_Systems ----
write('source_systems', sheet_to_rows('Source_Systems'))

# ---- Completeness ----
write('completeness', sheet_to_rows('Completeness'))

# ---- Dim_Area ----
write('dim_area', sheet_to_rows('Dim_Area'))

# ---- Dim_Discipline ----
write('dim_discipline', sheet_to_rows('Dim_Discipline'))

# ---- EVM_Quadrant ----
write('evm_quadrant', sheet_to_rows('EVM_Quadrant'))

# ---- Critical_Milestones ----
write('critical_milestones', sheet_to_rows('Critical_Milestones'))

# ---- Risk_Triggers ----
write('risk_triggers', sheet_to_rows('Risk_Triggers'))

# ---- Delay_Breakdown ----
write('delay_breakdown', sheet_to_rows('Delay_Breakdown'))

# ---- Decision_Log ----
write('decision_log', sheet_to_rows('Decision_Log'))

# ---- Why_Rate_Failing ----
if 'Why_Rate_Failing' in wb.sheetnames:
    write('why_rate_failing', sheet_to_rows('Why_Rate_Failing'))

# ---- Manpower ----
if 'Manpower' in wb.sheetnames:
    write('manpower', sheet_to_rows('Manpower'))

# ---- Hidden_Delay (single object, not array) ----
if 'Hidden_Delay' in wb.sheetnames:
    rows = sheet_to_rows('Hidden_Delay')
    write('hidden_delay', rows[0] if rows else {})

# ---- Recovery_Validation (single object, not array) ----
if 'Recovery_Validation' in wb.sheetnames:
    rows = sheet_to_rows('Recovery_Validation')
    write('recovery_validation', rows[0] if rows else {})

# ---- PPC_Trend ----
if 'PPC_Trend' in wb.sheetnames:
    write('ppc_trend', sheet_to_rows('PPC_Trend'))

# ---- Baseline_History ----
if 'Baseline_History' in wb.sheetnames:
    write('baseline_history', sheet_to_rows('Baseline_History'))

print('Done. ' + str(len(os.listdir(OUT))) + ' JSON files written to data/')
